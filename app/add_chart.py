import sys
import openpyxl
from openpyxl.chart import BarChart, Reference

file_path = sys.argv[1]
wb = openpyxl.load_workbook(file_path)

if "Dashboard" in wb.sheetnames:
    ws = wb["Dashboard"]

    # Find Monthly Performance data: header is in col 1 (A), row 14
    # Data starts row 15, find last row
    header_row = 14
    data_start = header_row + 1
    max_row = data_start - 1
    while ws.cell(row=max_row + 1, column=1).value is not None:
        max_row += 1

    if max_row >= data_start:
        chart = BarChart()
        chart.type    = "col"
        chart.style   = 10
        chart.title   = "Receipts and Payments by Month"
        chart.grouping = "clustered"

        # Receipts = col 2 (B), Payments = col 3 (C)
        data = Reference(ws, min_col=2, min_row=header_row, max_col=3, max_row=max_row)
        cats = Reference(ws, min_col=1,  min_row=data_start, max_row=max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Size: fit within right-panel columns
        chart.width  = 16   # cm
        chart.height = 8.5    # cm

        # Anchor: F14 (Col 6, Row 14)
        ws.add_chart(chart, "F14")

# Ensure calculation is set to auto so formulas evaluate
if wb.calculation:
    wb.calculation.calcMode    = "auto"
    wb.calculation.fullCalcOnLoad = True

wb.save(file_path)
